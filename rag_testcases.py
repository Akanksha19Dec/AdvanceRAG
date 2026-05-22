"""
RAG Application for Test Cases — ChromaDB + Nomic Embed Text (Ollama) + Groq LLM
=================================================================================
- Reads test cases from Test_Cases_VWO_Login.xlsx
- Each test case row becomes a chunk (natural chunking for structured data)
- Embeds using nomic-embed-text via local Ollama
- Stores in ChromaDB (local persistent)
- Answers user queries about test cases via Groq LLM
- Serves HTML UI with pipeline visualization and Q&A
"""

import math
import os
import requests
import chromadb
import openpyxl
from dotenv import load_dotenv
from flask import Flask, jsonify, request, send_from_directory

load_dotenv()

# ── Configuration ──
OLLAMA_BASE_URL = "http://localhost:11434"
EMBED_MODEL = "nomic-embed-text"

GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "openai/gpt-oss-120b"

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Test_Cases_VWO_Login.xlsx")
CHROMA_DB_PATH = os.path.join(os.path.dirname(__file__), "chroma_tc_db")
COLLECTION_NAME = "vwo_test_cases"
TOP_K = 5

app = Flask(__name__, static_folder=".", static_url_path="")


# ── Load Test Cases from Excel ──
def load_test_cases(path: str) -> list:
    """Load test cases from Excel and return as list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Test Cases"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        tc = dict(zip(headers, row))
        # Build a text representation for embedding
        text = (
            f"Test Case: {tc.get('Test Case ID','')}\n"
            f"Module: {tc.get('Module','')}\n"
            f"Category: {tc.get('Category','')}\n"
            f"Description: {tc.get('Description','')}\n"
            f"Pre-conditions: {tc.get('Pre-conditions','')}\n"
            f"Steps: {tc.get('Steps','')}\n"
            f"Expected Result: {tc.get('Expected Result','')}\n"
            f"Priority: {tc.get('Priority','')}"
        )
        cases.append({
            "id": len(cases),
            "tc_id": str(tc.get("Test Case ID", "")),
            "module": str(tc.get("Module", "")),
            "category": str(tc.get("Category", "")),
            "description": str(tc.get("Description", "")),
            "preconditions": str(tc.get("Pre-conditions", "")),
            "steps": str(tc.get("Steps", "")),
            "expected": str(tc.get("Expected Result", "")),
            "priority": str(tc.get("Priority", "")),
            "text": text,
            "length": len(text),
        })
    wb.close()
    return cases


# ── Embedding ──
def get_embedding(text: str) -> list:
    resp = requests.post(
        f"{OLLAMA_BASE_URL}/api/embeddings",
        json={"model": EMBED_MODEL, "prompt": text},
        timeout=60,
    )
    resp.raise_for_status()
    return resp.json()["embedding"]


# ── ChromaDB ──
def init_chromadb():
    return chromadb.PersistentClient(path=CHROMA_DB_PATH)


def populate_chromadb(client, cases: list):
    existing = [c.name for c in client.list_collections()]
    if COLLECTION_NAME in existing:
        col = client.get_collection(name=COLLECTION_NAME)
        if col.count() == len(cases):
            print(f"   -> Collection already has {col.count()} items. Skipping.")
            return col
        client.delete_collection(name=COLLECTION_NAME)

    col = client.create_collection(name=COLLECTION_NAME, metadata={"hnsw:space": "cosine"})
    print("   -> Embedding test cases and storing in ChromaDB...")
    for i, tc in enumerate(cases):
        emb = get_embedding(tc["text"])
        col.add(
            ids=[f"tc_{tc['id']}"],
            embeddings=[emb],
            documents=[tc["text"]],
            metadatas=[{
                "chunk_id": tc["id"], "tc_id": tc["tc_id"],
                "module": tc["module"], "category": tc["category"],
                "priority": tc["priority"], "length": tc["length"],
            }],
        )
        print(f"   -> Embedded {i+1}/{len(cases)}: {tc['tc_id']}")
    return col


def query_chromadb(col, query: str, top_k: int = TOP_K):
    emb = get_embedding(query)
    return col.query(query_embeddings=[emb], n_results=top_k,
                     include=["documents", "metadatas", "distances"])


def query_chromadb_by_module(col, module_name: str):
    """Retrieve ALL test cases for a specific module using metadata filter."""
    results = col.get(
        where={"module": module_name},
        include=["documents", "metadatas"],
    )
    return results


# ── Smart Module Detection ──
MODULE_KEYWORDS = {
    "Login Authentication": ["login", "authentication", "auth", "sign in", "signin"],
    "Password Management": ["password", "forgot password", "reset password", "password management"],
    "Session Management": ["session", "timeout", "logout", "cookie", "remember me"],
    "Input Validation": ["input validation", "validation", "email format", "real-time validation"],
    "SSO Integration": ["sso", "oauth", "saml", "google login", "microsoft login", "social login", "single sign"],
    "UI/UX Design": ["ui", "ux", "responsive", "mobile", "dark mode", "light mode", "design", "branding"],
    "Accessibility": ["accessibility", "wcag", "screen reader", "aria", "keyboard navigation", "a11y"],
    "Security": ["security", "https", "brute force", "gdpr", "compliance", "encryption", "tls"],
}


def detect_module(query: str) -> str | None:
    """Detect if a query is asking about a specific module."""
    q = query.lower()
    best_match = None
    best_score = 0
    for module, keywords in MODULE_KEYWORDS.items():
        for kw in keywords:
            if kw in q:
                score = len(kw)  # Longer keyword = more specific match
                if score > best_score:
                    best_score = score
                    best_match = module
    return best_match


# ── Groq LLM ──
def generate_answer(query: str, retrieved: list, stats_context: str = "") -> str:
    context = "\n\n---\n\n".join(
        [f"[{c['tc_id']}]:\n{c['text']}" for c in retrieved]
    )
    system_msg = (
        "You are a Senior QA Lead. Answer questions about VWO Login Dashboard test cases "
        "based on the provided context. Always reference specific Test Case IDs and list ALL "
        "relevant ones from the context. If asked about counts, list every matching test case ID.\n\n"
        f"DATASET OVERVIEW:\n{stats_context}"
    )
    payload = {
        "model": GROQ_MODEL,
        "messages": [
            {"role": "system", "content": system_msg},
            {"role": "user", "content": f"CONTEXT (Retrieved Test Cases):\n{context}\n\nQUESTION: {query}\n\nList ALL relevant Test Case IDs from the context."},
        ],
        "temperature": 0.3, "max_tokens": 2048,
    }
    resp = requests.post(GROQ_API_URL, headers={
        "Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"
    }, json=payload, timeout=60)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


# ── Startup ──
print("=" * 60)
print("  Test Case RAG — ChromaDB + Nomic Embed + Groq")
print("=" * 60)

print("\n[1/3] Loading test cases from Excel...")
test_cases = load_test_cases(EXCEL_PATH)
print(f"   -> {len(test_cases)} test cases loaded")

print("\n[2/3] Initializing ChromaDB...")
chroma_client = init_chromadb()
collection = populate_chromadb(chroma_client, test_cases)
print(f"   -> ChromaDB ready. Count: {collection.count()}")

print("\n[3/3] Checking Groq API...")
if GROQ_API_KEY:
    print(f"   -> Groq key loaded (model: {GROQ_MODEL})")
else:
    print("   ⚠️  GROQ_API_KEY not set in .env!")

# Build stats context string for LLM
module_stats = {}
category_stats = {}
priority_stats = {}
for tc in test_cases:
    module_stats[tc["module"]] = module_stats.get(tc["module"], 0) + 1
    category_stats[tc["category"]] = category_stats.get(tc["category"], 0) + 1
    priority_stats[tc["priority"]] = priority_stats.get(tc["priority"], 0) + 1

stats_context = f"Total test cases: {len(test_cases)}\n"
stats_context += "Modules: " + ", ".join(f"{m} ({c})" for m, c in module_stats.items()) + "\n"
stats_context += "Categories: " + ", ".join(f"{c} ({n})" for c, n in category_stats.items()) + "\n"
stats_context += "Priorities: " + ", ".join(f"{p} ({n})" for p, n in priority_stats.items())

print("\n[OK] Test Case RAG ready!\n")


# ── Routes ──
@app.route("/")
def index():
    return send_from_directory(".", "testcases.html")


@app.route("/api/chunks")
def api_chunks():
    all_data = collection.get(
        ids=[f"tc_{tc['id']}" for tc in test_cases], include=["embeddings"]
    )
    enriched = []
    has_emb = all_data["embeddings"] is not None and len(all_data["embeddings"]) > 0
    for i, tc in enumerate(test_cases):
        emb = list(all_data["embeddings"][i]) if has_emb else []
        emb_info = {}
        if len(emb) > 0:
            n = len(emb)
            emb_info = {
                "dimension": n,
                "magnitude": round(math.sqrt(sum(x*x for x in emb)), 4),
                "min": round(min(emb), 6), "max": round(max(emb), 6),
                "mean": round(sum(emb)/n, 6),
                "preview": [round(v, 6) for v in emb[:6]],
            }
        enriched.append({**tc, "embedding": emb_info})

    # Module/category/priority stats
    modules = {}
    cats = {}
    pris = {}
    for tc in test_cases:
        modules[tc["module"]] = modules.get(tc["module"], 0) + 1
        cats[tc["category"]] = cats.get(tc["category"], 0) + 1
        pris[tc["priority"]] = pris.get(tc["priority"], 0) + 1

    return jsonify({
        "total": len(test_cases),
        "embed_model": EMBED_MODEL, "llm_model": GROQ_MODEL,
        "llm_provider": "Groq", "vector_store": "ChromaDB",
        "embed_dim": enriched[0]["embedding"].get("dimension", 0) if enriched else 0,
        "modules": modules, "categories": cats, "priorities": pris,
        "chunks": enriched,
    })


@app.route("/api/ask", methods=["POST"])
def api_ask():
    data = request.get_json()
    query = data.get("question", "").strip()
    if not query:
        return jsonify({"error": "Please provide a question."}), 400
    if not GROQ_API_KEY:
        return jsonify({"error": "GROQ_API_KEY not set in .env"}), 500
    try:
        retrieved = []
        detected_module = detect_module(query)

        if detected_module:
            # Module-specific query: retrieve ALL test cases from that module
            results = query_chromadb_by_module(collection, detected_module)
            for i in range(len(results["ids"])):
                meta = results["metadatas"][i]
                retrieved.append({
                    "chunk_id": meta["chunk_id"], "tc_id": meta["tc_id"],
                    "module": meta["module"], "category": meta["category"],
                    "priority": meta["priority"],
                    "similarity": 1.0,  # Exact metadata match
                    "text": results["documents"][i],
                    "length": meta["length"],
                })
        else:
            # General query: use semantic similarity search
            results = query_chromadb(collection, query, TOP_K)
            for i in range(len(results["ids"][0])):
                meta = results["metadatas"][0][i]
                dist = results["distances"][0][i]
                retrieved.append({
                    "chunk_id": meta["chunk_id"], "tc_id": meta["tc_id"],
                    "module": meta["module"], "category": meta["category"],
                    "priority": meta["priority"],
                    "similarity": round(1 - dist, 4),
                    "text": results["documents"][0][i],
                    "length": meta["length"],
                })

        answer = generate_answer(query, retrieved, stats_context)
        return jsonify({"answer": answer, "retrieved_chunks": retrieved})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    print("[START] Test Case RAG at http://localhost:5001")
    app.run(debug=False, port=5001)
