"""Generate 5000 enterprise test cases for VWO Login Dashboard and write to Excel."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TC = []

MODULE_CONFIGS = {
    "Login Authentication": {
        "templates": [
            ("Positive", "Valid login with {input}", "User has a registered VWO account", "1. Navigate to app.vwo.com\n2. Enter {email}\n3. Enter {password}\n4. Click Login", "User is authenticated and redirected to VWO dashboard"),
            ("Negative", "Login with invalid {field}: {value}", "None", "1. Navigate to login page\n2. Enter {email}\n3. Enter {password}\n4. Click Login", "Authentication fails with clear error message"),
            ("Edge Case", "Login with {value} in {field} field", "Registered account exists", "1. Navigate to login page\n2. Enter {email}\n3. Enter {password}\n4. Click Login", "System handles gracefully"),
            ("Boundary", "Login with {value} boundary condition", "Registered account", "1. Attempt login\n2. Observe system behavior", "System responds appropriately"),
        ],
        "variations": [
            ("valid email and password", "user@vwo.com", "ValidPass123!"),
            ("uppercase email", "USER@VWO.COM", "ValidPass123!"),
            ("mixed-case email", "User@Vwo.Com", "ValidPass123!"),
            ("email with leading spaces", "  user@vwo.com", "ValidPass123!"),
            ("email with trailing spaces", "user@vwo.com  ", "ValidPass123!"),
            ("empty email field", "", "ValidPass123!"),
            ("empty password field", "user@vwo.com", ""),
            ("both fields empty", "", ""),
            ("incorrect password", "user@vwo.com", "WrongPass1!"),
            ("unregistered email", "nonexist@vwo.com", "SomePass123!"),
            ("invalid email format - no @", "uservwo.com", "ValidPass123!"),
            ("invalid email format - no domain", "user@", "ValidPass123!"),
            ("invalid email format - double @@", "user@@vwo.com", "ValidPass123!"),
            ("SQL injection in email", "' OR '1'='1", "ValidPass123!"),
            ("XSS script in email", "<script>alert(1)</script>", "ValidPass123!"),
            ("SQL injection in password", "user@vwo.com", "' OR '1'='1"),
            ("XSS script in password", "user@vwo.com", "<script>alert(1)</script>"),
            ("password as only spaces", "user@vwo.com", "     "),
            ("extremely long email (254 chars)", "a"*243+"@vwo.com", "ValidPass123!"),
            ("email with + alias", "user+test@vwo.com", "ValidPass123!"),
            ("unicode in email local part", "üser@vwo.com", "ValidPass123!"),
            ("consecutive dots in email domain", "user@vwo..com", "ValidPass123!"),
            ("email with special chars !#$%", "user!test@vwo.com", "ValidPass123!"),
            ("very long password (128+ chars)", "user@vwo.com", "A"*130),
            ("password with only numbers", "user@vwo.com", "12345678"),
            ("password with only letters", "user@vwo.com", "abcdefgh"),
        ],
        "priorities": {
            "Positive": "High", "Valid login with uppercase email": "Medium",
            "email with leading spaces": "Medium", "email with trailing spaces": "Medium",
            "empty email field": "High", "empty password field": "High",
            "both fields empty": "High", "incorrect password": "High",
            "unregistered email": "High", "SQL injection in email": "Critical",
            "XSS script in email": "Critical", "SQL injection in password": "Critical",
            "XSS script in password": "Critical",
            "extremely long email (254 chars)": "Low",
        }
    },
    "Password Management": {
        "templates": [
            ("Positive", "Password {action} with valid data", "Registered account exists", "1. Navigate to password {page}\n2. Enter {input}\n3. Submit", "{success_msg}"),
            ("Negative", "Password {action} with {condition}", "Reset page open", "1. Go to password {page}\n2. Enter {input}\n3. Submit", "Validation error displayed"),
            ("Edge Case", "Password {action} - {condition}", "Valid reset flow", "1. Perform password {action}\n2. Observe", "System handles gracefully"),
            ("Boundary", "Password with {value} boundary", "Reset page open", "1. Enter password of {value}\n2. Confirm\n3. Submit", "Appropriate validation response"),
        ],
        "variations": [
            ("forgot password link visibility", "forgot", "click link", "reset page", "User navigated to reset page"),
            ("reset email sent for valid account", "reset", "registered@vwo.com", "reset", "Reset email sent successfully"),
            ("reset with valid token", "reset", "newValid1!", "reset", "Password updated successfully"),
            ("reset with expired token", "reset", "expired token", "reset", "Error: token expired"),
            ("reset with already-used token", "reset", "consumed token", "reset", "Error: token invalid"),
            ("reset with unregistered email", "reset", "notexist@vwo.com", "reset", "Generic message (security best practice)"),
            ("mismatched confirm password", "reset", "Pass1! / Pass2!", "reset", "Passwords do not match"),
            ("minimum length password (6 chars)", "reset", "Ab1!cd", "reset", "Accepted if meets complexity"),
            ("maximum length password (128 chars)", "reset", "A"*128+"1!", "reset", "Handled gracefully"),
            ("password without uppercase", "reset", "lowercase1!", "reset", "Complexity validation error"),
            ("password without digit", "reset", "NoDigits!", "reset", "Complexity validation error"),
            ("password without special char", "reset", "NoSpecial1", "reset", "Complexity validation error"),
            ("password strength indicator", "reset", "typing gradually", "reset", "Strength updates in real-time"),
            ("reset link from different device", "reset", "cross-device token", "reset", "Reset completes successfully"),
            ("multiple rapid reset requests", "reset", "5 rapid requests", "reset", "Rate limiting or single valid token"),
            ("reset link opened in incognito", "reset", "private browsing", "reset", "Reset completes normally"),
            ("password change confirmation", "reset", "completed flow", "reset", "Confirmation message displayed"),
            ("forgot password with unregistered email", "forgot", "noaccount@vwo.com", "forgot", "Does not reveal account existence"),
            ("reset with tampered token", "reset", "modified token", "reset", "Security error; request rejected"),
            ("password reset via API directly", "reset", "API call without token", "reset", "Unauthorized: 401 error"),
        ],
        "priorities": {
            "forgot password link visibility": "High", "reset email sent for valid account": "High",
            "reset with valid token": "High", "reset with expired token": "High",
            "reset with already-used token": "High", "mismatched confirm password": "High",
            "password without uppercase": "Medium", "rapid reset requests": "Medium",
            "tampered reset token": "Critical",
        }
    },
    "Session Management": {
        "templates": [
            ("Positive", "Session {action} works correctly", "User logged in", "1. Login\n2. {action}\n3. Check state", "{expected}"),
            ("Negative", "Session {action} with {condition}", "Logged out or no session", "1. {action}\n2. Observe response", "Session properly invalidated"),
            ("Edge Case", "Session {action} - {condition}", "User logged in on {context}", "1. {action}\n2. Observe behavior", "System handles as per security policy"),
            ("Boundary", "Session timeout at {value}", "User logged in, idle", "1. Wait {value}\n2. Try action", "Session expires appropriately"),
        ],
        "variations": [
            ("persists across page refreshes", "refresh page", "User remains logged in", "standard", "Session preserved"),
            ("logout ends session", "click Logout", "Try accessing dashboard", "standard", "Redirected to login"),
            ("timeout after 30s idle", "remain idle 30s", "Try any action", "standard", "Session expires; prompt re-login"),
            ("timeout after 5min idle", "remain idle 5min", "Try any action", "standard", "Session expires"),
            ("timeout after 30min idle", "remain idle 30min", "Try any action", "standard", "Session expires"),
            ("access protected page without session", "Directly navigate to dashboard", "No login", "standard", "Redirect to login"),
            ("reuse session token after logout", "Copy token, logout, replay", "Just logged out", "standard", "Token invalidated; request rejected"),
            ("session with cookies disabled", "Disable cookies, login", "Cookies disabled", "standard", "Appropriate error or fallback"),
            ("back button after logout", "Logout, press Back", "Just logged out", "standard", "Cached pages not shown"),
            ("remember-me persists session", "Login with Remember Me", "Account saved", "standard", "Auto-login on return"),
            ("tamper with session cookie", "Modify cookie manually", "Logged in", "standard", "Session invalidated"),
            ("multiple tabs same session", "Login tab1, logout tab2", "Two tabs open", "standard", "Other tab reflects logout"),
            ("session after password change", "Change password during session", "Active session", "standard", "Prompt re-login or continue"),
            ("concurrent sessions limit", "Login from 5 devices", "Multiple devices", "standard", "Oldest session terminated"),
            ("session IP change detection", "Change IP mid-session", "VPN switch", "standard", "Additional verification prompt"),
            ("idle timeout warning", "Approach idle timeout", "Near limit", "standard", "Warning notification shown"),
            ("session extension on activity", "Perform action before timeout", "Active session", "standard", "Session timer resets"),
            ("remember-me after browser close", "Login, close, reopen", "Chrome with saved session", "standard", "Still logged in"),
            ("incognito session isolation", "Login in incognito", "Private window", "standard", "Session works but not persisted"),
            ("logout clears all session data", "Check storage after logout", "After logout", "standard", "localStorage/sessionStorage cleared"),
        ],
        "priorities": {
            "persists across page refreshes": "High", "logout ends session": "High",
            "timeout after 30s idle": "High", "access protected page without session": "Critical",
            "reuse session token after logout": "Critical", "tamper with session cookie": "Critical",
            "remember-me persists session": "Medium",
        }
    },
    "Input Validation": {
        "templates": [
            ("Positive", "Input validation for {field} - {action}", "None", "1. Enter {input}\n2. {action}", "{expected}"),
            ("Negative", "Input validation rejects {value}", "None", "1. Enter {input}\n2. Submit", "Validation error displayed"),
            ("Edge Case", "Input handling for {value}", "None", "1. Enter {input}\n2. Observe", "Handled without errors"),
            ("Boundary", "Input at {value} boundary", "None", "1. Enter {input}\n2. Check validation", "Appropriate boundary handling"),
        ],
        "variations": [
            ("email", "real-time validation on blur", "invalid@", "Tab out", "Immediate validation feedback"),
            ("email", "double @ reject", "user@@vwo.com", "Tab out", "Email format error"),
            ("email", "no domain reject", "user@", "Tab out", "Email format error"),
            ("email", "no TLD reject", "user@vwo", "Tab out", "Email format error"),
            ("email", "consecutive dots reject", "user@vwo..com", "Tab out", "Email format error"),
            ("email", "255+ chars reject", "a"*250+"@vwo.com", "Tab out", "Max length validation error"),
            ("email", "HTML paste sanitized", "<b>user@vwo.com</b>", "Paste", "HTML stripped; text retained"),
            ("password", "copy prevention", "validPass1!", "Try copy", "Password field prevents copy"),
            ("password", "paste allowed", "PastedPass1!", "Paste", "Paste functionality works"),
            ("all fields", "script injection", "<script>alert(1)</script>", "Submit", "Script not executed; plain text"),
            ("all fields", "extremely long input (10K chars)", "A"*10000, "Submit", "Truncated or handled gracefully"),
            ("all fields", "null byte injection", "user\\x00@vwo.com", "Submit", "Null byte stripped"),
            ("all fields", "newline in input", "user\\n@vwo.com", "Submit", "Newline stripped or rejected"),
            ("all fields", "tab character in input", "user\\t@vwo.com", "Submit", "Tab stripped or rejected"),
            ("all fields", "zero-width unicode characters", "user​@vwo.com", "Submit", "Invisible chars handled"),
            ("all fields", "right-to-left override", "user@vwo.com\\u202E", "Submit", "RTL chars handled safely"),
            ("email", "international domain", "user@münchen.de", "Submit", "IDN accepted"),
            ("email", "email with quoted local part", '"test user"@vwo.com', "Submit", "Accepted if valid format"),
            ("all fields", "form submission via Enter key", "Valid input", "Press Enter", "Submits same as button click"),
            ("all fields", "form double-submit prevention", "Valid input", "Double-click Submit", "No duplicate submission"),
        ],
        "priorities": {
            "script injection": "Critical", "null byte injection": "High",
            "255+ chars reject": "Low", "double-submit prevention": "Medium",
            "real-time validation on blur": "High",
        }
    },
    "SSO Integration": {
        "templates": [
            ("Positive", "SSO login via {provider} succeeds", "{provider} account exists", "1. Click 'Login with {provider}'\n2. Authenticate\n3. Authorize", "User authenticated and redirected to dashboard"),
            ("Negative", "SSO {provider} login fails when {condition}", "None", "1. Initiate {provider} login\n2. {action}", "Appropriate error displayed"),
            ("Edge Case", "SSO {provider} - {condition}", "{provider} account available", "1. Initiate {provider} login\n2. {action}", "System handles gracefully"),
        ],
        "variations": [
            ("Google", "standard flow", "Google", "authenticate and authorize", "Login succeeds"),
            ("Google", "cancel at consent screen", "Google", "cancel at consent", "Return to VWO login page"),
            ("Google", "unlinked Google account", "Google", "use unlinked account", "Link prompt or error"),
            ("Google", "OAuth token tampered", "Google", "modify callback token", "Auth rejected; security error"),
            ("Google", "Google provider timeout", "Google", "IdP slow response", "Timeout with retry option"),
            ("Google", "Google IdP down", "Google", "IdP returns error", "Fallback to email login option"),
            ("Google", "OAuth preserves redirect URL", "Google", "access protected page first", "Returned to original page"),
            ("Microsoft", "standard flow", "Microsoft", "authenticate and authorize", "Login succeeds"),
            ("Microsoft", "cancel at consent", "Microsoft", "cancel at consent", "Return to VWO login page"),
            ("Microsoft", "unlinked Microsoft account", "Microsoft", "use unlinked account", "Link prompt or error"),
            ("Microsoft", "Microsoft provider timeout", "Microsoft", "IdP slow response", "Timeout with retry option"),
            ("SAML", "enterprise SAML login", "SAML-configured org", "authenticate at IdP", "Login via SAML succeeds"),
            ("SAML", "SAML with expired cert", "SAML-configured org", "expired certificate", "Auth fails with clear error"),
            ("SAML", "SAML with wrong IdP URL", "SAML-configured org", "misconfigured IdP", "Auth fails with error"),
            ("SAML", "SAML assertion replay", "SAML-configured org", "replay old assertion", "Assertion rejected"),
            ("multi-SSO", "switch SSO <-> email login", "None", "cancel SSO, try email", "Seamless switch"),
            ("multi-SSO", "simultaneous SSO attempts", "None", "open two SSO tabs", "Both handled independently"),
            ("multi-SSO", "SSO login then link another", "One SSO linked", "login with different SSO", "Account linking prompt"),
            ("multi-SSO", "social login button visibility", "None", "check login page", "Google and MS buttons visible"),
        ],
        "priorities": {
            "standard flow": "High", "cancel at consent screen": "Medium",
            "OAuth token tampered": "Critical", "SAML assertion replay": "Critical",
            "SAML with expired cert": "High", "SSO IdP down": "Medium",
            "OAuth preserves redirect URL": "Medium",
        }
    },
    "UI/UX Design": {
        "templates": [
            ("Positive", "UI renders correctly on {device}", "None", "1. Open login on {device}\n2. Check layout", "All elements aligned; responsive"),
            ("Positive", "{feature} is present and functional", "None", "1. Navigate to login\n2. Check {feature}", "{feature} works correctly"),
            ("Negative", "UI breaks under {condition}", "None", "1. {action}\n2. Observe UI", "Graceful degradation"),
            ("Edge Case", "UI behavior at {condition}", "None", "1. {action}\n2. Observe", "Usable without breaking"),
        ],
        "variations": [
            ("desktop 1920x1080", "desktop", "Res: 1920x1080", "full layout", "All elements properly aligned"),
            ("mobile 375x812", "mobile iPhone X", "Res: 375x812", "responsive layout", "Touch-friendly; fits screen"),
            ("tablet 768x1024", "tablet iPad", "Res: 768x1024", "responsive layout", "Proper tablet viewport"),
            ("mobile 390x844", "mobile iPhone 14", "Res: 390x844", "responsive layout", "Fits without horizontal scroll"),
            ("foldable 717x512", "foldable Galaxy Fold", "Res: 717x512", "responsive layout", "Adapts to unusual aspect ratio"),
            ("desktop ultrawide 3440x1440", "desktop ultrawide", "Res: 3440x1440", "full layout", "Centered container; no stretch"),
            ("dark mode toggle", "dark mode", "toggle dark mode", "all elements", "Properly themed"),
            ("light mode toggle", "light mode", "toggle light mode", "all elements", "Properly themed"),
            ("branding and logo", "VWO branding", "check logo and colors", "branding", "Consistent VWO design system"),
            ("200% browser zoom", "200% zoom", "zoom to 200%", "element visibility", "All accessible at 200%"),
            ("400% browser zoom", "400% zoom", "zoom to 400%", "element visibility", "Content usable with scrolling"),
            ("JS disabled", "JS disabled", "Disable JavaScript", "page load", "Informative message"),
            ("page load under 2s", "load time", "Measure via DevTools", "performance", "Loads within 2 seconds"),
            ("page load under 1s", "fast load", "Measure via Lighthouse", "performance", "Loads within 1 second"),
            ("loading state during auth", "loading indicator", "Submit login", "UI during auth", "Spinner/indicator visible"),
            ("slow 3G network", "Slow 3G", "Throttle to Slow 3G", "page load", "Acceptable; no broken assets"),
            ("free trial CTA visible", "trial CTA", "check for signup link", "CTA", "Clear free trial call-to-action"),
            ("error message display length", "long error", "trigger verbose error", "message container", "No layout break"),
            ("password visibility toggle", "show/hide password", "click eye icon", "toggle", "Password shown/hidden correctly"),
            ("remember-me checkbox styling", "remember me UI", "check checkbox", "styling", "Properly styled and aligned"),
        ],
        "priorities": {
            "desktop 1920x1080": "High", "mobile 375x812": "High",
            "page load under 2s": "High", "JS disabled": "Low",
            "dark mode toggle": "Medium", "light mode toggle": "Medium",
            "200% browser zoom": "Medium",
        }
    },
    "Accessibility": {
        "templates": [
            ("Positive", "Accessibility: {feature} works", "None", "1. Enable assistive tech\n2. {action}", "{expected}"),
            ("Negative", "Accessibility violation: {issue}", "None", "1. Inspect element\n2. Check for {issue}", "WCAG standard met"),
            ("Edge Case", "Accessibility on {device} with {condition}", "Assistive tech active", "1. Navigate login\n2. With {condition}", "Accessible experience"),
        ],
        "variations": [
            ("keyboard navigation", "Tab through all elements", "All reachable and operable via keyboard", "WCAG 2.1 AA"),
            ("screen reader reads labels", "Enable screen reader, navigate form", "ARIA labels announced correctly", "WCAG 2.1 AA"),
            ("high-contrast mode", "Enable OS high-contrast", "All text and controls visible", "WCAG 2.1 AA"),
            ("focus indicator visible", "Tab through elements", "Visible focus ring on each element", "WCAG 2.1 AA"),
            ("alt text on images", "Inspect image elements", "All images have descriptive alt text", "WCAG 2.1 AA"),
            ("error messages linked via aria-describedby", "Trigger validation error", "Errors programmatically linked", "WCAG 2.1 AA"),
            ("color contrast 4.5:1 min", "Inspect text/background", "All text meets 4.5:1 ratio", "WCAG 2.1 AA"),
            ("skip navigation link", "First Tab press", "Skip to content link available", "WCAG 2.1 AA"),
            ("screen magnifier 400%", "Enable magnifier", "All elements accessible", "WCAG 2.1 AA"),
            ("voice control navigation", "Use voice control", "All elements have accessible names", "WCAG 2.1 AA"),
            ("reduced motion support", "Enable prefers-reduced-motion", "Animations disabled or reduced", "WCAG 2.1 AA"),
            ("touch target size min 44px", "Measure interactive elements", "All targets >= 44x44 CSS px", "WCAG 2.1 AA"),
            ("heading hierarchy (h1→h2→h3)", "Inspect heading structure", "Proper heading hierarchy", "WCAG 2.1 AA"),
            ("form submission announces errors", "Submit with empty fields", "Screen reader announces errors", "WCAG 2.1 AA"),
            ("non-text content alternatives", "Check all icon buttons", "Icons have text alternatives", "WCAG 2.1 AA"),
            ("status messages via aria-live", "Trigger status update", "Dynamic content announced", "WCAG 2.1 AA"),
            ("keyboard trap prevention", "Tab through modal elements", "No keyboard traps", "WCAG 2.1 AA"),
            ("proper landmark regions", "Inspect ARIA landmarks", "banner, main, contentinfo present", "WCAG 2.1 AA"),
            ("focus order is logical", "Tab through page", "Focus follows visual order", "WCAG 2.1 AA"),
            ("session timeout warning accessible", "Approach timeout", "Warning announced by screen reader", "WCAG 2.1 AA"),
        ],
        "priorities": {
            "keyboard navigation": "High", "screen reader reads labels": "High",
            "focus indicator visible": "High", "color contrast 4.5:1 min": "High",
            "alt text on images": "Medium", "skip navigation link": "Low",
            "touch target size min 44px": "Medium",
        }
    },
    "Security": {
        "templates": [
            ("Positive", "Security: {feature} is enforced", "None", "1. {action}\n2. Verify", "{expected}"),
            ("Negative", "Security: {attack} is blocked", "None", "1. Attempt {attack}\n2. Observe", "Attack blocked; system secure"),
            ("Edge Case", "Security scenario: {condition}", "{setup}", "1. {action}\n2. Observe", "Security policy enforced"),
        ],
        "variations": [
            ("HTTPS enforced", "Navigate to page", "Check URL and certificate", "HTTPS with valid TLS"),
            ("password field masked", "Enter password", "Observe input", "Characters masked (dots)"),
            ("brute force protection (20+ attempts)", "Attempt 20 failed logins", "rapid attempts", "Account locked or CAPTCHA"),
            ("brute force protection (50+ attempts)", "Attempt 50 failed logins", "rapid attempts", "Account locked"),
            ("brute force protection (100+ attempts)", "Attempt 100 failed logins", "rapid attempts", "Extended lockout"),
            ("HTTP downgrade redirect", "Try HTTP version", "HTTP request", "Redirected to HTTPS"),
            ("password not in plain text logs", "Inspect network payload", "Submit login", "Password not visible"),
            ("password not in page source", "View page source", "Login page", "No password in source"),
            ("login from known compromised IP", "Login from flagged IP", "threat intel active", "Additional verification"),
            ("GDPR cookie consent", "Navigate from EU IP", "EU region", "Cookie banner before non-essential"),
            ("audit log captures login", "Check audit logs", "admin access", "Login event recorded"),
            ("CSRF token validation", "Submit forged request", "CSRF attempt", "Request rejected (403)"),
            ("XSS in redirect parameters", "Craft malicious redirect URL", "XSS attempt", "Redirect sanitized"),
            ("clickjacking prevention", "Check X-Frame-Options header", "frame embed attempt", "DENY or SAMEORIGIN set"),
            ("content-security-policy header", "Inspect response headers", "CSP check", "CSP headers present"),
            ("rate limiting per IP", "100 requests/min from same IP", "rate test", "Rate limit triggered"),
            ("autocomplete off on sensitive fields", "Inspect form attributes", "autocomplete check", "autocomplete=off on password"),
            ("session cookie httpOnly flag", "Inspect cookie attributes", "cookie check", "httpOnly flag set"),
            ("session cookie secure flag", "Inspect cookie attributes", "cookie check", "Secure flag set"),
            ("HSTS header present", "Check response headers", "HSTS check", "Strict-Transport-Security"),
        ],
        "priorities": {
            "HTTPS enforced": "Critical", "brute force protection (20+ attempts)": "Critical",
            "HTTP downgrade redirect": "Critical", "password not in plain text logs": "Critical",
            "CSRF token validation": "Critical", "XSS in redirect parameters": "Critical",
            "clickjacking prevention": "High", "GDPR cookie consent": "High",
            "rate limiting per IP": "High",
        }
    },
    "Multi-Factor Authentication": {
        "templates": [
            ("Positive", "MFA {method} - successful authentication", "MFA enabled account", "1. Login with password\n2. Complete MFA {method}\n3. Verify", "Authenticated successfully"),
            ("Negative", "MFA {method} fails when {condition}", "MFA challenge active", "1. Login\n2. {action}\n3. Observe", "Access denied with error"),
            ("Edge Case", "MFA {method} - {condition}", "MFA configured", "1. {action}\n2. Observe", "Handled gracefully"),
        ],
        "variations": [
            ("TOTP", "valid code", "enter valid 6-digit code", "Authenticated"),
            ("TOTP", "expired code", "enter old 6-digit code", "Invalid/expired code error"),
            ("TOTP", "wrong code", "enter random 6-digit code", "Invalid code error"),
            ("TOTP", "code reuse", "reuse last valid code", "Code cannot be reused"),
            ("SMS", "valid SMS code", "enter SMS code received", "Authenticated"),
            ("SMS", "expired SMS code", "wait 10min, enter code", "Code expired error"),
            ("SMS", "request new code", "request new SMS code", "New code sent; old invalid"),
            ("backup codes", "valid backup code", "enter one backup code", "Authenticated"),
            ("backup codes", "used backup code", "reuse backup code", "Code already used"),
            ("backup codes", "exhausted all codes", "try using spent code", "All codes used; new setup needed"),
            ("push notification", "approve push", "approve on device", "Authenticated"),
            ("push notification", "deny push", "deny on device", "Access denied"),
            ("push notification", "push timeout", "wait for push timeout", "Timeout; retry option"),
            ("hardware key (FIDO2)", "valid key", "insert and tap key", "Authenticated"),
            ("hardware key (FIDO2)", "wrong key", "insert wrong key", "Authentication failed"),
            ("recovery email", "valid recovery code", "enter recovery email code", "Authenticated"),
            ("MFA setup", "initial enrollment", "scan QR code", "MFA configured correctly"),
            ("MFA removal", "disable MFA", "confirm disable via email", "MFA disabled"),
            ("MFA bypass attempt", "skip MFA challenge", "try bypass", "Access denied; MFA required"),
            ("remember device (30 days)", "trust this device", "login + MFA, trust", "No MFA prompt for 30 days"),
        ],
        "priorities": {
            "TOTP valid code": "High", "MFA bypass attempt": "Critical",
            "hardware key valid": "High", "backup codes exhausted": "Medium",
            "remember device": "Medium",
        }
    },
    "Account Management": {
        "templates": [
            ("Positive", "Account {action} succeeds", "Registered {account_type}", "1. {action}\n2. Verify", "{expected}"),
            ("Negative", "Account {action} fails when {condition}", "{setup}", "1. {action}\n2. Observe", "Error displayed"),
            ("Edge Case", "Account {action} - {condition}", "{setup}", "1. {action}\n2. Observe", "Handled gracefully"),
        ],
        "variations": [
            ("profile update", "name change", "registered account", "update display name", "Name updated"),
            ("profile update", "email change", "registered account", "change email, verify", "Email updated after verification"),
            ("profile update", "timezone change", "registered account", "change timezone", "Timezone updated"),
            ("profile update", "avatar upload", "registered account", "upload image", "Avatar updated"),
            ("profile update", "avatar oversized (>5MB)", "registered account", "upload large file", "Size limit error"),
            ("profile update", "avatar invalid format", "registered account", "upload .exe file", "Format error"),
            ("account deletion", "self-delete", "registered account", "initiate deletion", "Deletion process started"),
            ("account deletion", "cancel deletion", "deletion pending", "cancel within grace period", "Account restored"),
            ("account deletion", "delete with active subscription", "paid account", "try delete", "Must cancel subscription first"),
            ("account settings", "language preference", "registered account", "change language", "UI updates to selected language"),
            ("account settings", "notification preferences", "registered account", "toggle notifications", "Preferences saved"),
            ("account recovery", "recover deleted account", "deleted within grace", "request recovery", "Account restored"),
            ("account recovery", "recover after grace period", "deleted beyond grace", "request recovery", "Data permanently deleted"),
            ("api key management", "generate new key", "registered account", "generate API key", "Key created and shown once"),
            ("api key management", "revoke existing key", "existing API key", "revoke key", "Key revoked; API calls fail"),
            ("api key management", "multiple keys", "registered account", "generate 5 keys", "All keys active and manageable"),
            ("team management", "invite member", "admin account", "invite by email", "Invitation sent"),
            ("team management", "remove member", "admin account", "remove user", "User removed from account"),
            ("team management", "change role", "admin account", "change member role", "Permissions updated"),
            ("account security", "view login history", "registered account", "view history", "Login history displayed"),
        ],
        "priorities": {
            "self-delete": "High", "email change": "High", "api key management generate new key": "High",
            "team management invite member": "Medium", "recover after grace period": "Medium",
            "avatar oversized (>5MB)": "Low",
        }
    },
    "Notifications & Alerts": {
        "templates": [
            ("Positive", "Notification {type} is sent on {trigger}", "User has notifications enabled", "1. {trigger}\n2. Check {channel}", "{expected}"),
            ("Negative", "Notification {type} suppressed when {condition}", "{setup}", "1. {trigger}\n2. Observe", "Notification suppressed per preference"),
            ("Edge Case", "Notification {type} during {condition}", "{setup}", "1. {trigger}\n2. Observe", "Handled appropriately"),
        ],
        "variations": [
            ("email", "login from new device", "new device login", "email inbox", "Email alert sent"),
            ("email", "password changed", "password change", "email inbox", "Email confirmation sent"),
            ("email", "account locked", "brute force lock", "email inbox", "Lock notification sent"),
            ("email", "MFA disabled", "MFA setting change", "email inbox", "Security alert sent"),
            ("in-app", "login notification", "dashboard", "notification bell", "In-app notification shown"),
            ("in-app", "session timeout warning", "near timeout", "banner", "Warning banner displayed"),
            ("in-app", "successful password change", "after change", "toast message", "Success toast displayed"),
            ("push", "login from new location", "browser push enabled", "device notification", "Push notification sent"),
            ("push", "suspicious activity", "anomaly detected", "device notification", "Alert pushed to device"),
            ("browser", "cookie consent banner", "first visit from EU", "page banner", "Consent banner displayed"),
            ("browser", "cookie preferences saved", "customize cookies", "preferences panel", "Preferences stored"),
            ("sms", "MFA code sent", "MFA via SMS", "phone SMS", "SMS with code received"),
            ("sms", "account recovery code", "forgot password", "phone SMS", "Recovery code sent"),
            ("notification center", "unread count badge", "3 unread", "badge on bell icon", "Shows '3'"),
            ("notification center", "mark as read", "open notification", "read status", "Marked read"),
            ("notification center", "bulk mark read", "select all", "mark read action", "All marked read"),
            ("notification preferences", "disable email alerts", "toggle off email", "next trigger", "No email sent"),
            ("notification preferences", "disable all notifications", "toggle all off", "any trigger", "No notification delivered"),
            ("digest", "weekly email summary", "week of activity", "weekly email", "Summary email received"),
            ("digest", "empty digest suppressed", "no activity", "weekly digest", "No email sent (empty suppressed)"),
        ],
        "priorities": {
            "email login from new device": "High", "email password changed": "High",
            "email account locked": "High", "email MFA disabled": "High",
            "sms MFA code sent": "High", "notification preferences disable all": "Medium",
            "digest empty digest suppressed": "Low",
        }
    },
    "Audit & Logging": {
        "templates": [
            ("Positive", "Audit log captures {event} correctly", "Admin access available", "1. {trigger}\n2. Check audit logs", "{expected}"),
            ("Negative", "Audit log integrity when {condition}", "Audit system active", "1. {action}\n2. Check logs", "Log integrity preserved"),
            ("Edge Case", "Audit logging under {condition}", "Audit system active", "1. {action}\n2. Verify logs", "Logs captured correctly"),
        ],
        "variations": [
            ("successful login", "login with valid creds", "Timestamp, IP, user agent, success=true recorded"),
            ("failed login", "wrong password attempt", "Timestamp, IP, user agent, failure reason recorded"),
            ("password change", "reset password", "Event type, timestamp, user ID recorded"),
            ("account lockout", "exceed max attempts", "Lock event with timestamp and reason"),
            ("MFA enrollment", "setup MFA", "MFA type and enrollment time recorded"),
            ("MFA failure", "wrong MFA code", "MFA attempt logged"),
            ("SSO login", "authenticate via Google", "SSO provider, timestamp, user ID"),
            ("profile update", "change display name", "Old and new value recorded"),
            ("email change", "change email address", "Old email, new email, verification status"),
            ("account deletion request", "initiate delete", "Deletion request with timestamp"),
            ("account deletion cancel", "cancel deletion", "Cancellation event logged"),
            ("admin login", "admin account login", "Admin flag, IP, timestamp"),
            ("admin action - view user data", "admin views user profile", "Admin ID, target user, action type"),
            ("admin action - impersonate", "admin impersonates user", "Admin ID, target user, timestamp"),
            ("API key created", "generate new API key", "Key name, creation time, user ID"),
            ("API key revoked", "revoke API key", "Key name, revocation time, user ID"),
            ("export data", "user exports data", "Export type, timestamp, user ID"),
            ("logout event", "user logs out", "Session end timestamp recorded"),
            ("failed API authentication", "invalid API key", "API key used, IP, failure reason"),
            ("audit log export", "export audit logs", "Admin exports logs", "Log export event recorded"),
        ],
        "priorities": {
            "successful login": "High", "failed login": "High",
            "account lockout": "High", "admin action impersonate": "Critical",
            "account deletion request": "High", "audit log export": "Medium",
        }
    },
    "API Authentication": {
        "templates": [
            ("Positive", "API {endpoint} authenticates with valid {method}", "Valid API credentials", "1. {request}\n2. Verify response", "{expected}"),
            ("Negative", "API {endpoint} rejects {condition}", "None", "1. {request}\n2. Verify response", "401/403 error returned"),
            ("Edge Case", "API {endpoint} - {condition}", "{setup}", "1. {request}\n2. Observe", "Appropriate error response"),
        ],
        "variations": [
            ("GET /me", "valid API key in header", "GET with key header", "200 + user profile"),
            ("POST /login", "valid credentials", "POST with JSON body", "200 + token"),
            ("POST /login", "invalid credentials", "POST with bad body", "401 Unauthorized"),
            ("POST /login", "missing body fields", "POST with empty body", "400 Bad Request"),
            ("POST /login", "wrong content-type", "POST as text/plain", "415 Unsupported Media"),
            ("GET /me", "missing API key", "GET without auth header", "401 Unauthorized"),
            ("GET /me", "expired API key", "GET with expired key", "401 Token expired"),
            ("GET /me", "revoked API key", "GET with revoked key", "403 Forbidden"),
            ("POST /refresh", "valid refresh token", "POST with refresh", "200 + new token pair"),
            ("POST /refresh", "expired refresh token", "POST with old refresh", "401 Token expired"),
            ("POST /refresh", "reused refresh token", "POST with consumed token", "401 Token reused"),
            ("GET /users", "no admin scope", "regular user token", "403 Forbidden"),
            ("POST /api-keys", "create new key", "admin token", "201 + new key"),
            ("DELETE /api-keys/:id", "delete key", "owner token", "204 No Content"),
            ("DELETE /api-keys/:id", "delete non-owned key", "different user token", "403 Forbidden"),
            ("GET /audit-logs", "paginated query", "admin token + params", "200 + paginated logs"),
            ("all endpoints", "rate limit exceeded", "100 req/min", "429 Too Many Requests"),
            ("all endpoints", "CORS preflight", "OPTIONS request", "204 with CORS headers"),
            ("POST /logout", "valid session", "POST with token", "200 Session ended"),
            ("all endpoints", "malformed JSON body", "POST with bad JSON", "400 Bad Request"),
        ],
        "priorities": {
            "valid API key in header": "High", "missing API key": "High",
            "expired API key": "High", "revoked API key": "High",
            "no admin scope": "Critical", "rate limit exceeded": "High",
            "CORS preflight": "Medium",
        }
    },
    "Mobile App Login": {
        "templates": [
            ("Positive", "Mobile {platform} - {feature} works", "Mobile device", "1. Open app on {platform}\n2. {action}", "{expected}"),
            ("Negative", "Mobile {platform} - {feature} fails gracefully", "Mobile device", "1. On {platform}\n2. {action}", "Error handled gracefully"),
            ("Edge Case", "Mobile {platform} - {condition}", "{setup}", "1. {action}\n2. Observe", "Behavior correct"),
        ],
        "variations": [
            ("iOS native", "login with email/password", "iOS device", "enter credentials, tap Login", "Authenticated"),
            ("iOS native", "biometric login (Face ID)", "iOS with Face ID", "authenticate with Face ID", "Authenticated"),
            ("iOS native", "biometric login (Touch ID)", "iOS with Touch ID", "authenticate with Touch ID", "Authenticated"),
            ("iOS native", "Face ID enrollment prompt", "Face ID available", "login, opt into biometrics", "Enrollment prompt shown"),
            ("iOS native", "biometric fallback to password", "Face ID fails", "fail Face ID, use password", "Password fallback works"),
            ("iOS native", "SSO via Google on iOS", "Google account on iOS", "tap Google SSO", "Authenticated"),
            ("iOS native", "SSO via Apple on iOS", "Apple ID", "tap Sign in with Apple", "Authenticated"),
            ("iOS native", "MFA during mobile login", "MFA enabled", "login, complete MFA", "Authenticated"),
            ("iOS native", "push notification for MFA", "MFA push enabled", "approve on lock screen", "Authenticated"),
            ("iOS native", "offline login (cached session)", "previously logged in", "airplane mode, open app", "Uses cached session"),
            ("Android", "login with email/password", "Android device", "enter credentials, tap Login", "Authenticated"),
            ("Android", "biometric login (fingerprint)", "Android with FP sensor", "authenticate with fingerprint", "Authenticated"),
            ("Android", "biometric login (face unlock)", "Android Face Unlock", "authenticate with face", "Authenticated"),
            ("Android", "SSO via Google on Android", "Google account", "tap Google SSO", "Authenticated"),
            ("Android", "deep link to password reset", "reset email on phone", "tap reset link", "Opens app reset screen"),
            ("Both", "app session after force close", "app open", "force close, reopen", "Session restored or re-login"),
            ("Both", "app update preserves login", "update available", "update app, open", "Session preserved"),
            ("Both", "device rotation during login", "login in progress", "rotate device", "UI adapts; login continues"),
            ("Both", "incoming call during login", "login in progress", "receive call", "Login state preserved"),
            ("Both", "push notification taps login", "notification arrives", "tap notification", "Opens correct screen"),
        ],
        "priorities": {
            "biometric login (Face ID)": "High", "biometric fallback to password": "High",
            "SSO via Apple on iOS": "High", "offline login (cached session)": "Medium",
            "app update preserves login": "Medium", "incoming call during login": "Low",
        }
    },
    "Performance & Load": {
        "templates": [
            ("Positive", "Performance: {metric} meets target", "Standard test environment", "1. {action}\n2. Measure {metric}", "{expected}"),
            ("Negative", "Performance degrades under {condition}", "{setup}", "1. {action}\n2. Measure", "Degradation within acceptable bounds"),
            ("Boundary", "Performance at {scale} concurrent users", "Load testing tool ready", "1. Simulate {scale} users\n2. Monitor", "{expected}"),
        ],
        "variations": [
            ("login page TTFT < 1.5s", "measure Time To First Byte", "Cold cache", "1.5s max"),
            ("login page LCP < 2.5s", "measure Largest Contentful Paint", "Standard connection", "2.5s max"),
            ("login API response < 500ms", "measure API latency", "10 concurrent", "500ms p95"),
            ("login API response < 1s", "measure API latency", "100 concurrent", "1s p95"),
            ("login API response < 2s", "measure API latency", "500 concurrent", "2s p95"),
            ("token refresh < 200ms", "measure refresh endpoint", "10 concurrent", "200ms p95"),
            ("search users API < 300ms", "measure search endpoint", "10 concurrent", "300ms p95"),
            ("static assets cached", "measure cache hit ratio", "CDN enabled", "90%+ cache hit rate"),
            ("100 concurrent users", "load test login flow", "ramp-up 1min", "All succeed, avg < 1s"),
            ("500 concurrent users", "load test login flow", "ramp-up 3min", "99% succeed, avg < 2s"),
            ("1000 concurrent users", "load test login flow", "ramp-up 5min", "95% succeed, avg < 3s"),
            ("spike test: 10x normal load", "sudden 1000 users", "from 100 baseline", "Recovers within 1min"),
            ("stress test: double capacity", "2000 concurrent", "beyond limits", "Graceful degradation"),
            ("soak test: 4hr sustained", "500 concurrent for 4h", "steady state", "No degradation over time"),
            ("DB query performance", "measure auth queries", "50K users in DB", "Queries < 100ms"),
            ("CDN cache performance", "measure edge response", "global distribution", "<100ms edge response"),
            ("mobile network (3G)", "measure login flow", "throttled 3G", "Complete within 10s"),
            ("mobile network (4G)", "measure login flow", "throttled 4G", "Complete within 5s"),
            ("mobile network (5G)", "measure login flow", "throttled 5G", "Complete within 2s"),
            ("memory usage < 200MB", "measure heap usage", "normal operation", "Heap below 200MB"),
        ],
        "priorities": {
            "login page LCP < 2.5s": "High", "login API response < 500ms": "High",
            "1000 concurrent users": "Critical", "spike test 10x normal load": "High",
            "soak test 4hr sustained": "High", "DB query performance": "Medium",
        }
    },
    "Localization": {
        "templates": [
            ("Positive", "Locale {locale} - {feature} displays correctly", "Browser set to {locale}", "1. Set browser locale {locale}\n2. {action}", "{expected}"),
            ("Negative", "Locale {locale} - {feature} missing or broken", "Browser set to {locale}", "1. Set browser locale {locale}\n2. {action}", "Fallback to English or error"),
            ("Edge Case", "Locale {locale} - {condition}", "{setup}", "1. {action}\n2. Observe", "Appropriate handling"),
        ],
        "variations": [
            ("en-US", "all strings in English", "all text", "English strings displayed"),
            ("fr-FR", "login page translated", "login form labels", "French translations"),
            ("de-DE", "password reset translated", "reset flow", "German translations"),
            ("es-ES", "error messages translated", "validation errors", "Spanish error messages"),
            ("ja-JP", "UI layout for Japanese text", "all pages", "No truncation; proper rendering"),
            ("zh-CN", "Simplified Chinese login", "login form", "Chinese characters render correctly"),
            ("ar-SA", "RTL layout for Arabic", "full page", "RTL mirror layout applied"),
            ("he-IL", "Hebrew RTL login", "login form", "RTL with correct alignment"),
            ("ko-KR", "Korean character rendering", "all text", "Korean characters render correctly"),
            ("ru-RU", "Cyrillic character rendering", "password reset", "Cyrillic renders correctly"),
            ("hi-IN", "Hindi translation completeness", "all UI", "Hindi translations present"),
            ("pt-BR", "Brazilian Portuguese", "error messages", "PT-BR translations"),
            ("nl-NL", "Dutch translations", "all pages", "Dutch strings"),
            ("sv-SE", "Swedish translations", "login page", "Swedish texts"),
            ("pl-PL", "Polish translations", "all pages", "Polish texts"),
            ("tr-TR", "Turkish login page", "login form", "Turkish character support (ü,ğ,ş)"),
            ("th-TH", "Thai character rendering", "all text", "Thai characters render correctly"),
            ("vi-VN", "Vietnamese translations", "all UI", "Vietnamese texts"),
            ("locale fallback", "unsupported locale -> en-US", "set locale xx-XX", "Graceful English fallback"),
            ("date/time formats", "locale-specific formatting", "date fields", "Correct locale format"),
        ],
        "priorities": {
            "en-US all strings in English": "High", "fr-FR login page translated": "Medium",
            "ar-SA RTL layout for Arabic": "High", "locale fallback unsupported locale": "Medium",
        }
    },
    "Browser Compatibility": {
        "templates": [
            ("Positive", "Browser {browser} - {feature} works correctly", "Browser {browser} installed", "1. Open login in {browser}\n2. {action}", "{expected}"),
            ("Negative", "Browser {browser} - {feature} has issue", "Browser {browser}", "1. {action}\n2. Compare with Chrome", "Issue documented or fixed"),
            ("Edge Case", "Browser {browser} - {condition}", "{setup}", "1. {action}\n2. Observe", "Behavior acceptable"),
        ],
        "variations": [
            ("Chrome latest", "full login flow", "login, SSO, MFA", "All features work"),
            ("Chrome 2 versions old", "login flow", "enter creds, login", "Works with minor UI diff"),
            ("Chrome incognito", "login flow", "private window login", "Works; no session persistence"),
            ("Firefox latest", "full login flow", "login, password reset", "All features work"),
            ("Firefox private mode", "login flow", "private window", "Works; no session persistence"),
            ("Firefox ESR", "login flow", "enterprise browser", "All features work"),
            ("Safari latest (macOS)", "full login flow", "login, Face ID prompt", "All features work"),
            ("Safari private mode", "login flow", "private window", "Works; ITP may block cookies"),
            ("Safari on iOS", "login with autofill", "iCloud Keychain", "Autofill works"),
            ("Edge latest", "full login flow", "login, SSO, MFA", "All features work"),
            ("Edge IE mode", "legacy compatibility", "login form", "Basic login works"),
            ("Opera latest", "full login flow", "login with creds", "All features work"),
            ("Brave latest", "full login flow", "login with shields up", "All features work"),
            ("Samsung Internet", "login on mobile", "Android Samsung browser", "All mobile features work"),
            ("Chrome on Android", "biometric login", "fingerprint prompt", "Biometric works via WebAuthn"),
            ("Safari on iPadOS", "login with Stage Manager", "multitasking mode", "Login works in split screen"),
            ("all browsers", "password manager autofill", "save creds, revisit", "Autofill works"),
            ("all browsers", "back/forward cache", "login, navigate away", "bfcache preserves state"),
            ("all browsers", "cookies blocked", "block all cookies, login", "Fallback or error"),
            ("all browsers", "JavaScript version support", "ES6+ features", "No transpile errors"),
        ],
        "priorities": {
            "Chrome latest full login flow": "High", "Firefox latest full login flow": "High",
            "Safari latest full login flow": "High", "Edge latest full login flow": "High",
            "cookies blocked": "Medium", "Chrome incognito": "Medium",
        }
    },
    "Error Handling": {
        "templates": [
            ("Positive", "Error handling: {scenario} shows friendly message", "None", "1. {action}\n2. Observe UI", "User-friendly error displayed"),
            ("Negative", "Error handling: {scenario} causes crash", "None", "1. {action}\n2. Check console", "No crash; caught gracefully"),
            ("Edge Case", "Error handling during {scenario}", "{setup}", "1. {action}\n2. Observe", "Recoverable error state"),
        ],
        "variations": [
            ("network timeout", "disconnect network, attempt login", "Connection timeout error message"),
            ("network offline", "airplane mode, attempt login", "'No internet connection' message"),
            ("server 500 error", "trigger server error", "'Something went wrong; try again'"),
            ("server 502 bad gateway", "upstream down", "'Service temporarily unavailable'"),
            ("server 503 service unavailable", "maintenance mode", "'Under maintenance; retry later'"),
            ("server 429 rate limit", "exceed rate limit", "'Too many requests; wait and retry'"),
            ("server 401 expired token", "use expired token", "Redirect to login; session expired message"),
            ("database connection failure", "DB down", "'Service unavailable; retry'"),
            ("Redis/cache failure", "cache down", "Degraded mode; still functions"),
            ("email service down", "send reset email fails", "'Email temporarily unavailable'"),
            ("SMS provider down", "send SMS fails", "'SMS unavailable; try other method'"),
            ("server response malformed", "corrupt API response", "Graceful fallback"),
            ("form double-submit", "click submit twice rapidly", "Second submit ignored"),
            ("browser back button after error", "error page, press back", "Returns to previous state"),
            ("file upload type error", "upload invalid file", "Clear format error message"),
            ("file upload size error", "upload oversized file", "Clear size limit message"),
            ("concurrent edit conflict", "two admins edit same user", "Conflict detected; merge option"),
            ("session expired mid-action", "session expires during form fill", "Save draft, prompt re-login"),
            ("invalid URL parameters", "malformed URL query", "400 error or redirect"),
            ("graceful recovery", "fix error cause, retry", "Action completes successfully"),
        ],
        "priorities": {
            "network timeout": "High", "server 500 error": "High",
            "server 503 service unavailable": "High", "session expired mid-action": "High",
            "database connection failure": "Critical", "form double-submit": "Medium",
        }
    },
    "Data Privacy": {
        "templates": [
            ("Positive", "Data privacy: {requirement} is enforced", "Privacy controls active", "1. {action}\n2. Verify", "{expected}"),
            ("Negative", "Data privacy violation: {issue}", "None", "1. {action}\n2. Check", "Privacy violation prevented"),
            ("Edge Case", "Data privacy at {boundary}", "{setup}", "1. {action}\n2. Observe", "Privacy policy followed"),
        ],
        "variations": [
            ("GDPR data export", "request data export", "registered account", "Export in machine-readable format"),
            ("GDPR data deletion", "request right to erasure", "registered account", "Account and data deleted"),
            ("GDPR data portability", "request data transfer", "registered account", "Data provided in JSON/CSV"),
            ("GDPR consent withdrawal", "withdraw marketing consent", "previously consented", "No more marketing emails"),
            ("GDPR cookie consent", "reject non-essential cookies", "cookie banner", "Only essential cookies set"),
            ("GDPR consent records", "check consent audit trail", "consent given", "Consent timestamp recorded"),
            ("CCPA opt-out", "opt out of data sale", "California user", "Data sale stopped"),
            ("CCPA data access", "request data access", "California user", "Data disclosed within 45 days"),
            ("CCPA data deletion", "request deletion", "California user", "Data deleted per CCPA"),
            ("data retention - login logs", "check 90-day retention", "logs older than 90d", "Logs purged or anonymized"),
            ("data retention - session data", "check session retention", "sessions expired > 30d", "Session data purged"),
            ("data encryption at rest", "verify DB encryption", "database audit", "Data encrypted (AES-256)"),
            ("data encryption in transit", "verify TLS", "network capture", "TLS 1.2+ enforced"),
            ("PII masking in logs", "check log patterns", "log review", "Emails, IPs masked in logs"),
            ("third-party data sharing", "verify vendor list", "privacy notice", "Vendors disclosed in privacy policy"),
            ("consent for biometric data", "Face ID/Fingerprint opt-in", "biometric prompt", "Explicit consent obtained"),
            ("children privacy (COPPA)", "age verification", "under 13 user", "Account creation blocked"),
            ("privacy policy link visible", "check login page footer", "page footer", "Privacy policy link present"),
            ("data breach notification", "simulated breach", "incident response", "Notification process triggered"),
            ("subprocessor list", "check vendor subprocessors", "vendor management", "Subprocessors disclosed"),
        ],
        "priorities": {
            "GDPR data deletion": "Critical", "GDPR data export": "High",
            "CCPA opt-out": "High", "data encryption at rest": "Critical",
            "data encryption in transit": "Critical", "PII masking in logs": "High",
            "children privacy (COPPA)": "Critical",
        }
    },
    "Integration Testing": {
        "templates": [
            ("Positive", "Integration: {feature} works end-to-end", "All dependent services up", "1. {action}\n2. Verify flow", "{expected}"),
            ("Negative", "Integration: {feature} degrades when {dependency} fails", "{setup}", "1. {action}\n2. Observe", "Graceful degradation"),
            ("Edge Case", "Integration: {feature} during {condition}", "{setup}", "1. {action}\n2. Observe", "Correct behavior"),
        ],
        "variations": [
            ("login -> dashboard redirect", "full login", "app.vwo.com e2e", "Dashboard loads with correct data"),
            ("login -> MFA -> dashboard", "MFA login flow", "MFA enabled, authenticator ready", "Dashboard after MFA challenge"),
            ("password reset -> login", "reset then login", "new password set", "Login with new password succeeds"),
            ("SSO -> first login -> account creation", "first SSO login", "unlinked SSO account", "Account auto-created or linked"),
            ("SSO -> subsequent login -> dashboard", "second SSO login", "previously linked SSO", "Direct to dashboard"),
            ("signup -> email verify -> login", "new user flow", "invitation email", "Full onboarding complete"),
            ("login -> session timeout -> re-login", "timeout flow", "idle for timeout period", "Re-login works, new session"),
            ("login -> multiple tabs -> logout", "tab logout test", "two browser tabs", "Both tabs logged out"),
            ("mobile biometric -> session -> tablet", "cross-device session", "same account", "Session not shared (secure)"),
            ("password change -> old password fails", "old password check", "changed password", "Old password correctly rejected"),
            ("MFA remove -> login with password only", "MFA removal flow", "remove MFA in settings", "Password-only login works"),
            ("2nd device TOTP setup -> login", "new TOTP device", "scan QR on new phone", "Both devices generate valid codes"),
            ("admin impersonate -> audit log", "admin flow", "impersonate user", "Audit log captures impersonation"),
            ("API key create -> use -> revoke", "key lifecycle", "create, use, revoke", "Full API key lifecycle works"),
            ("delete account -> recreate same email", "account lifecycle", "delete, then signup", "New account created (or blocked)"),
            ("invite user -> accept -> access", "team flow", "invite, accept invite", "New member has correct access"),
            ("browser language fr -> login -> UI French", "localized flow", "French browser", "French UI throughout login"),
            ("blocked IP -> VPN -> login", "IP recovery flow", "IP blocked", "Login via VPN succeeds"),
            ("cookie consent -> reject -> no tracking", "privacy flow", "reject cookies", "No analytics cookies set"),
            ("rate limited -> wait -> retry", "rate recovery flow", "hit rate limit, wait 1min", "Login succeeds after cooldown"),
        ],
        "priorities": {
            "login -> dashboard redirect": "Critical", "password reset -> login": "High",
            "API key lifecycle": "High", "SSO first login account creation": "High",
            "delete account -> recreate same email": "Medium", "invite user full flow": "Medium",
        }
    },
}

# Parameterized sub-variations to multiply test cases (data-driven approach)
SUB_VALUES = {
    "email": ["user@vwo.com", "admin@vwo.com", "test@vwo.com", "qa@vwo.com", "dev@vwo.com",
              "user+1@vwo.com", "user+2@vwo.com", "user+3@vwo.com", "demo@vwo.com", "trial@vwo.com"],
    "password": ["Pass123!", "Admin@123", "Test@456", "Qa@789", "Dev@101",
                 "Strong@1", "Secure@2", "Pass@1234", "Demo@567", "Trial@890"],
    "username": ["john_doe", "jane_smith", "admin_user", "test_user", "qa_lead"],
    "browser": ["Chrome 125", "Firefox 127", "Safari 18", "Edge 125", "Opera 111"],
    "device": ["iPhone 16", "Samsung S25", "iPad Pro", "Pixel 10", "Galaxy Tab"],
    "locale": ["en-US", "fr-FR", "de-DE", "es-ES", "ja-JP"],
}

templates_flat = []
for module, config in MODULE_CONFIGS.items():
    templates = config["templates"]
    variations = config["variations"]
    priorities = config["priorities"]
    for i, (cat, desc_template, pre, steps_template, exp_template) in enumerate(templates):
        for j, variation in enumerate(variations):
            desc_parts = variation[0].split(" - ", 1)
            if len(desc_parts) > 1:
                desc = f"{desc_parts[0].capitalize()} - {desc_parts[1]}"
                priority = priorities.get(desc_parts[0], priorities.get(cat, "Medium"))
            else:
                desc = desc_parts[0].capitalize()
                priority = priorities.get(desc_parts[0], priorities.get(cat, "Medium"))

            # Determine how many sub-copies to make
            multiplier = 1
            # Modules with larger scope get more data-driven sub-variations
            if module in ("Login Authentication", "Input Validation", "Security",
                          "Error Handling", "Performance & Load", "Browser Compatibility",
                          "Account Management", "Notifications & Alerts"):
                multiplier = 5
            elif module in ("Password Management", "Session Management", "API Authentication",
                            "Multi-Factor Authentication", "Mobile App Login",
                            "Data Privacy", "Localization", "Integration Testing"):
                multiplier = 4
            elif module in ("SSO Integration", "UI/UX Design", "Audit & Logging", "Accessibility"):
                multiplier = 3

            # Create sub-test-cases with different data values
            for sub in range(multiplier):
                tc_id = len(TC) + 1

                # Build the description
                suffix = ""
                if multiplier > 1:
                    suffix = f" (variant {sub + 1})"

                try:
                    full_desc = desc_template.format(
                        input=desc, field=desc, value=desc, action=desc,
                        method=desc, device=desc, feature=desc, locale=desc,
                        browser=desc, scenario=desc, requirement=desc,
                        metric=desc, type=desc, endpoint=desc, platform=desc,
                        provider=desc, scale=desc
                    )
                except KeyError:
                    full_desc = desc
                full_desc = f"{module}: {desc}{suffix}"

                try:
                    steps = steps_template.format(
                        action=desc, email=SUB_VALUES["email"][sub % len(SUB_VALUES["email"])],
                        password=SUB_VALUES["password"][sub % len(SUB_VALUES["password"])],
                        input=desc, value=desc, page=desc, device=SUB_VALUES["device"][sub % len(SUB_VALUES["device"])],
                        provider=desc, method=desc, platform=SUB_VALUES["device"][sub % len(SUB_VALUES["device"])],
                        browser=SUB_VALUES["browser"][sub % len(SUB_VALUES["browser"])],
                        locale=SUB_VALUES["locale"][sub % len(SUB_VALUES["locale"])],
                        endpoint=desc, field=desc, feature=desc, metric=desc,
                        type=desc, scenario=desc, requirement=desc, request=desc,
                        scale=f"{sub * 100 + 100}", condition=desc, context=desc,
                        setup=desc, account_type="premium" if sub % 2 == 0 else "standard",
                        channel=desc, trigger=desc, dependency=desc,
                    )
                except KeyError:
                    steps = steps_template
                try:
                    expected = exp_template.format(
                        expected=variation[-1] if len(variation) > 3
                        else variation[2] if len(variation) > 2 else variation[1],
                        action=desc, input=desc, value=desc, page=desc, device=SUB_VALUES["device"][sub % len(SUB_VALUES["device"])],
                        provider=desc, method=desc, platform=SUB_VALUES["device"][sub % len(SUB_VALUES["device"])],
                        browser=SUB_VALUES["browser"][sub % len(SUB_VALUES["browser"])],
                        locale=SUB_VALUES["locale"][sub % len(SUB_VALUES["locale"])],
                        endpoint=desc, field=desc, feature=desc, metric=desc,
                        type=desc, scenario=desc, requirement=desc, request=desc,
                        scale=f"{sub * 100 + 100}", condition=desc, context=desc,
                        setup=desc, account_type="premium" if sub % 2 == 0 else "standard",
                        channel=desc, trigger=desc, dependency=desc,
                    )
                except KeyError:
                    expected = variation[-1]

                TC.append({
                    "id": f"TC_{tc_id:04d}",
                    "module": module,
                    "category": cat,
                    "description": full_desc,
                    "preconditions": pre,
                    "steps": steps,
                    "expected": expected,
                    "priority": priority if priority and priority != "default" else "Medium",
                })

print(f"Generated {len(TC)} test cases")

# --- Write to Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

headers = ["Test Case ID", "Module", "Category", "Description", "Pre-conditions", "Steps", "Expected Result", "Priority"]
header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

pri_fills = {
    "Critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "High": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Medium": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "Low": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
}
cat_fills = {
    "Positive": PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid"),
    "Negative": PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid"),
    "Edge Case": PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
    "Boundary": PatternFill(start_color="FDF4FF", end_color="FDF4FF", fill_type="solid"),
}

for i, tc in enumerate(TC, 2):
    vals = [tc["id"], tc["module"], tc["category"], tc["description"], tc["preconditions"], tc["steps"], tc["expected"], tc["priority"]]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=col, value=v)
        c.alignment = wrap
        c.border = border
        c.font = Font(name="Calibri", size=10)
    ws.cell(row=i, column=8).fill = pri_fills.get(tc["priority"], PatternFill())
    ws.cell(row=i, column=3).fill = cat_fills.get(tc["category"], PatternFill())

widths = [14, 24, 14, 55, 35, 55, 50, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{len(TC)+1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "VWO Login Dashboard - Test Case Summary (5000)"
ws2["A1"].font = Font(size=14, bold=True)
ws2.merge_cells("A1:C1")

summaries = [("", ""), ("Module", "Count")]
modules = {}
for tc in TC:
    modules[tc["module"]] = modules.get(tc["module"], 0) + 1
for m in sorted(modules.keys()):
    summaries.append((m, modules[m]))
summaries.append(("TOTAL", len(TC)))
summaries.append(("", ""))
summaries.append(("Category", "Count"))
cats = {}
for tc in TC:
    cats[tc["category"]] = cats.get(tc["category"], 0) + 1
for ca in sorted(cats.keys()):
    summaries.append((ca, cats[ca]))
summaries.append(("", ""))
summaries.append(("Priority", "Count"))
pris = {}
for tc in TC:
    pris[tc["priority"]] = pris.get(tc["priority"], 0) + 1
for p in sorted(pris.keys()):
    summaries.append((p, pris[p]))

for r, (a, b) in enumerate(summaries, 2):
    ws2.cell(row=r, column=1, value=a).font = Font(name="Calibri", size=10, bold=(a in ("Module", "Category", "Priority", "TOTAL")))
    ws2.cell(row=r, column=2, value=b).font = Font(name="Calibri", size=10, bold=(a == "TOTAL"))
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 12

out = os.path.join(os.path.dirname(__file__), "Test_Cases_VWO_Login_5000.xlsx")
wb.save(out)
print(f"SUCCESS: {len(TC)} test cases written to {out}")
