"""Generate 100 enterprise-level test cases for VWO Login Dashboard and write to Excel."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TC = []
def add(mid, mod, cat, desc, pre, steps, exp, pri):
    TC.append({"id":f"TC_{mid:03d}","module":mod,"category":cat,"description":desc,"preconditions":pre,"steps":steps,"expected":exp,"priority":pri})

# === MODULE 1: LOGIN AUTHENTICATION (TC_001 - TC_020) ===
add(1,"Login Authentication","Positive","Valid login with correct email and password","User has a registered VWO account","1. Navigate to app.vwo.com\n2. Enter valid email\n3. Enter valid password\n4. Click Login","User is authenticated and redirected to VWO dashboard","High")
add(2,"Login Authentication","Positive","Login is case-insensitive for email","Registered account exists","1. Navigate to login page\n2. Enter email in UPPERCASE\n3. Enter correct password\n4. Click Login","Login succeeds regardless of email case","Medium")
add(3,"Login Authentication","Positive","Auto-focus on first input field on page load","None","1. Navigate to app.vwo.com login page\n2. Observe cursor position","Email input field is auto-focused","Medium")
add(4,"Login Authentication","Positive","Visible loading state during authentication","Valid credentials available","1. Enter valid credentials\n2. Click Login\n3. Observe UI during auth process","Loading spinner/indicator is displayed during authentication","Medium")
add(5,"Login Authentication","Negative","Login with empty email and password fields","None","1. Navigate to login page\n2. Leave both fields empty\n3. Click Login","Validation error displayed for required fields","High")
add(6,"Login Authentication","Negative","Login with valid email but empty password","Registered account","1. Enter valid email\n2. Leave password empty\n3. Click Login","Validation error for password field","High")
add(7,"Login Authentication","Negative","Login with empty email but valid password","None","1. Leave email empty\n2. Enter a password\n3. Click Login","Validation error for email field","High")
add(8,"Login Authentication","Negative","Login with incorrect password","Registered account","1. Enter valid email\n2. Enter wrong password\n3. Click Login","Authentication fails with clear error message","High")
add(9,"Login Authentication","Negative","Login with unregistered email","None","1. Enter non-existent email\n2. Enter any password\n3. Click Login","Authentication fails with appropriate error message","High")
add(10,"Login Authentication","Negative","Login with invalid email format (no @ symbol)","None","1. Enter 'testuser.com'\n2. Enter password\n3. Click Login","Email format validation error displayed","High")
add(11,"Login Authentication","Negative","Login with SQL injection in email field","None","1. Enter \"admin' OR '1'='1\" in email\n2. Enter any password\n3. Click Login","Input is sanitized; login fails safely","Critical")
add(12,"Login Authentication","Negative","Login with XSS script in email field","None","1. Enter '<script>alert(1)</script>' in email\n2. Click Login","Script is not executed; input is sanitized","Critical")
add(13,"Login Authentication","Edge Case","Login with email containing leading/trailing spaces","Registered account","1. Enter email with spaces: '  user@vwo.com  '\n2. Enter correct password\n3. Click Login","Email is trimmed; login succeeds","Medium")
add(14,"Login Authentication","Edge Case","Login with maximum length email (254 chars)","Account with 254-char email","1. Enter 254-character valid email\n2. Enter correct password\n3. Click Login","Login succeeds if email is valid","Low")
add(15,"Login Authentication","Edge Case","Login with special characters in email local part","Account with email like user+test@vwo.com","1. Enter email with + symbol\n2. Enter correct password\n3. Click Login","Login succeeds for valid special-char emails","Medium")
add(16,"Login Authentication","Boundary","Rapid consecutive login attempts (rate limiting)","Valid account","1. Attempt login 10+ times rapidly with wrong password","Rate limiting kicks in; user is temporarily locked or throttled","Critical")
add(17,"Login Authentication","Edge Case","Login after session timeout","Previously logged-in session expired","1. Wait for session timeout (30 sec)\n2. Try to access dashboard\n3. Observe redirect","User is redirected to login page","High")
add(18,"Login Authentication","Negative","Login with password containing only spaces","Registered account","1. Enter valid email\n2. Enter spaces-only password\n3. Click Login","Validation error; login fails","Medium")
add(19,"Login Authentication","Edge Case","Concurrent login from two different browsers","Valid account","1. Login in Browser A\n2. Login in Browser B with same account\n3. Check both sessions","Behavior defined per security policy (Needs clarification)","High")
add(20,"Login Authentication","Positive","Clickable labels for input fields","None","1. Navigate to login page\n2. Click on the label text for email/password fields","Corresponding input field receives focus","Low")

# === MODULE 2: PASSWORD MANAGEMENT (TC_021 - TC_035) ===
add(21,"Password Management","Positive","Forgot password link navigates to reset page","None","1. Navigate to login page\n2. Click 'Forgot Password' link","User is navigated to password reset page","High")
add(22,"Password Management","Positive","Password reset email sent for valid email","Registered account","1. Go to Forgot Password page\n2. Enter registered email\n3. Submit","Password reset email is sent to the user","High")
add(23,"Password Management","Positive","Reset password with valid token and new password","Reset email received","1. Click reset link from email\n2. Enter new valid password\n3. Confirm password\n4. Submit","Password is updated; user can login with new password","High")
add(24,"Password Management","Positive","Password strength indicator displayed","None","1. Navigate to password reset page\n2. Start typing a password","Password strength indicator updates in real-time","Medium")
add(25,"Password Management","Negative","Reset password with expired token","Reset token older than expiry period","1. Click expired reset link\n2. Try to set new password","Error message indicating token has expired","High")
add(26,"Password Management","Negative","Reset password with already-used token","Token already consumed","1. Use a previously used reset link\n2. Try to set new password","Error indicating token is invalid/already used","High")
add(27,"Password Management","Negative","Submit forgot password with unregistered email","None","1. Enter non-existent email on reset page\n2. Submit","Appropriate message displayed (should not reveal email existence per security best practices)","High")
add(28,"Password Management","Negative","Password reset with mismatched confirm password","Reset page open","1. Enter new password\n2. Enter different confirm password\n3. Submit","Validation error: passwords do not match","High")
add(29,"Password Management","Boundary","Password with minimum allowed length","Reset page open","1. Enter password at minimum length boundary\n2. Confirm and submit","Password accepted if meets complexity rules (Needs clarification on exact min length)","Medium")
add(30,"Password Management","Boundary","Password with maximum allowed length","Reset page open","1. Enter extremely long password (e.g., 128+ chars)\n2. Submit","System handles gracefully — accepts or shows max length error","Medium")
add(31,"Password Management","Edge Case","Password reset link accessed from different device/IP","Reset email received","1. Request reset on Device A\n2. Open reset link on Device B\n3. Set new password","Reset completes successfully (token-based, not device-bound)","Medium")
add(32,"Password Management","Negative","Password without uppercase letter (complexity rule)","Reset page open","1. Enter password with only lowercase\n2. Submit","Validation error per password complexity rules (Needs clarification)","Medium")
add(33,"Password Management","Negative","Password without special character","Reset page open","1. Enter password with only alphanumeric\n2. Submit","Validation error per complexity rules (Needs clarification)","Medium")
add(34,"Password Management","Edge Case","Multiple rapid forgot-password requests","Valid account","1. Submit forgot password 5 times rapidly","System handles gracefully; rate limiting or single valid token","Medium")
add(35,"Password Management","Positive","Successful password change confirmation message","Valid reset flow completed","1. Complete password reset flow","Clear confirmation message displayed","Medium")

# === MODULE 3: SESSION MANAGEMENT (TC_036 - TC_045) ===
add(36,"Session Management","Positive","Session persists across page refreshes","User logged in","1. Login successfully\n2. Refresh the page","User remains logged in; session is preserved","High")
add(37,"Session Management","Positive","Logout successfully ends session","User logged in","1. Click Logout\n2. Try accessing dashboard URL directly","Session is terminated; user redirected to login","High")
add(38,"Session Management","Positive","Session timeout after configured period (30 sec)","User logged in, idle","1. Login and remain idle for 30 seconds\n2. Try any action","Session expires; user is prompted to re-login","High")
add(39,"Session Management","Negative","Access protected page without active session","No login","1. Directly navigate to VWO dashboard URL without logging in","User is redirected to login page","Critical")
add(40,"Session Management","Negative","Reuse session token after logout","Logged out","1. Copy session token before logout\n2. Logout\n3. Replay token in API call","Token is invalidated; request is rejected","Critical")
add(41,"Session Management","Edge Case","Session behavior when browser cookies are disabled","None","1. Disable cookies\n2. Attempt to login","Appropriate error or fallback behavior","Medium")
add(42,"Session Management","Edge Case","Back button after logout","Just logged out","1. Logout\n2. Press browser Back button","Cached pages are not shown; user stays on login page","High")
add(43,"Session Management","Positive","Remember-me option persists credentials","Valid account","1. Login with 'Remember Me' checked\n2. Close browser\n3. Reopen and navigate to app.vwo.com","User is auto-logged in or credentials pre-filled","Medium")
add(44,"Session Management","Negative","Tamper with session cookie value","User logged in","1. Modify session cookie value manually\n2. Refresh page","Session is invalidated; user redirected to login","Critical")
add(45,"Session Management","Edge Case","Multiple tabs with same session","User logged in","1. Open app in multiple tabs\n2. Logout in one tab\n3. Perform action in other tab","Other tabs reflect logout state on next action","Medium")

# === MODULE 4: INPUT VALIDATION (TC_046 - TC_058) ===
add(46,"Input Validation","Positive","Real-time email validation on blur","None","1. Type invalid email\n2. Tab out of email field","Immediate validation feedback shown on blur","High")
add(47,"Input Validation","Positive","Real-time password validation on blur","None","1. Type short password\n2. Tab out","Password validation feedback displayed on blur","High")
add(48,"Input Validation","Negative","Email with double @ symbol","None","1. Enter 'user@@vwo.com'\n2. Tab out","Email format validation error","Medium")
add(49,"Input Validation","Negative","Email with no domain","None","1. Enter 'user@'\n2. Tab out","Email format validation error","Medium")
add(50,"Input Validation","Negative","Email with consecutive dots in domain","None","1. Enter 'user@vwo..com'\n2. Tab out","Email format validation error","Medium")
add(51,"Input Validation","Boundary","Email at exactly 255 characters (over limit)","None","1. Enter 255-character email\n2. Tab out","Validation error: email exceeds max length","Low")
add(52,"Input Validation","Edge Case","Paste email with HTML tags","None","1. Paste '<b>user@vwo.com</b>' into email field","HTML is stripped/escaped; only text retained","Medium")
add(53,"Input Validation","Negative","Password field accepts copy-paste","None","1. Type password\n2. Try to copy from password field","Password field should prevent copying (Needs clarification)","Low")
add(54,"Input Validation","Positive","Mobile-optimized keyboard for email field","Mobile device","1. Open login on mobile\n2. Tap email field","Email-optimized keyboard shown (with @ key visible)","Medium")
add(55,"Input Validation","Edge Case","Unicode characters in email local part","None","1. Enter email with unicode: 'üser@vwo.com'\n2. Tab out","Validation handles unicode appropriately","Low")
add(56,"Input Validation","Negative","Script injection via password field","None","1. Enter '<script>alert(1)</script>' as password\n2. Submit","Script is not executed; treated as plain text","Critical")
add(57,"Input Validation","Positive","Clear error messages for validation failures","None","1. Submit form with invalid data\n2. Observe error messages","Specific, clear error messages per field","High")
add(58,"Input Validation","Edge Case","Form submission via Enter key","Credentials entered","1. Enter email and password\n2. Press Enter key","Form submits same as clicking Login button","Medium")

# === MODULE 5: SSO & SOCIAL LOGIN (TC_059 - TC_070) ===
add(59,"SSO Integration","Positive","Login via Google OAuth","Google account exists","1. Click 'Login with Google'\n2. Authenticate with Google\n3. Authorize VWO","User authenticated and redirected to VWO dashboard","High")
add(60,"SSO Integration","Positive","Login via Microsoft OAuth","Microsoft account exists","1. Click 'Login with Microsoft'\n2. Authenticate\n3. Authorize","User authenticated and redirected to dashboard","High")
add(61,"SSO Integration","Positive","Enterprise SAML SSO login","SAML configured for org","1. Enter enterprise email\n2. Redirected to IdP\n3. Authenticate at IdP","User authenticated via SAML and redirected to VWO","High")
add(62,"SSO Integration","Negative","Cancel Google OAuth flow midway","None","1. Click 'Login with Google'\n2. Cancel/deny at Google consent screen","User returned to VWO login page with appropriate message","Medium")
add(63,"SSO Integration","Negative","SSO login with unlinked account","Google account not linked to VWO","1. Click 'Login with Google'\n2. Use unlinked Google account","Appropriate error or account linking prompt","Medium")
add(64,"SSO Integration","Negative","SAML SSO with expired certificate","Expired SAML cert","1. Attempt SAML login","Authentication fails with clear error","High")
add(65,"SSO Integration","Edge Case","SSO provider timeout","Network issues to IdP","1. Click SSO login\n2. IdP takes too long to respond","Timeout error with retry option displayed","Medium")
add(66,"SSO Integration","Positive","OAuth login preserves redirect URL","User was on specific page","1. Access protected page\n2. Redirected to login\n3. Login via Google OAuth","After OAuth, user returned to original requested page","Medium")
add(67,"SSO Integration","Negative","Tampered OAuth callback token","None","1. Initiate OAuth\n2. Modify callback token parameter\n3. Complete flow","Authentication rejected; security error displayed","Critical")
add(68,"SSO Integration","Edge Case","SSO login when provider is down","IdP unavailable","1. Click SSO login\n2. Provider returns error","User shown error with option to use email/password login","Medium")
add(69,"SSO Integration","Positive","Social login button visibility on login page","None","1. Navigate to login page\n2. Observe social login options","Google and Microsoft login buttons are visible and clickable","Medium")
add(70,"SSO Integration","Edge Case","Switch between SSO and email login","None","1. Click Google login\n2. Cancel\n3. Login with email/password","Seamless switch between auth methods","Low")

# === MODULE 6: UI/UX & RESPONSIVE DESIGN (TC_071 - TC_082) ===
add(71,"UI/UX Design","Positive","Login page renders correctly on desktop (1920x1080)","None","1. Open login page on desktop browser\n2. Check layout and elements","All elements properly aligned; no overflow or cutoff","High")
add(72,"UI/UX Design","Positive","Login page renders on mobile (375x812)","Mobile device","1. Open login page on mobile\n2. Check responsive layout","Page is fully responsive; touch-friendly controls","High")
add(73,"UI/UX Design","Positive","Login page renders on tablet (768x1024)","Tablet device","1. Open login page on tablet\n2. Check layout","Proper responsive design for tablet viewport","Medium")
add(74,"UI/UX Design","Positive","Light and Dark mode toggle","None","1. Toggle between Light and Dark mode\n2. Observe all elements","All elements properly themed in both modes","Medium")
add(75,"UI/UX Design","Negative","Login page with JavaScript disabled","None","1. Disable JS in browser\n2. Navigate to login page","Graceful degradation or informative message","Low")
add(76,"UI/UX Design","Positive","VWO branding elements present","None","1. Navigate to login page\n2. Check logo, colors, fonts","VWO design system branding is consistent","Medium")
add(77,"UI/UX Design","Edge Case","Login page zoom at 200%","None","1. Zoom browser to 200%\n2. Check all elements","All elements accessible and usable at 200% zoom","Medium")
add(78,"UI/UX Design","Edge Case","Very long error message display","Error triggered","1. Trigger a validation error\n2. Check message container","Error message contained properly; no layout break","Low")
add(79,"UI/UX Design","Positive","Page load time under 2 seconds","Standard connection","1. Navigate to login page\n2. Measure load time via DevTools","Login page loads within 2 seconds","High")
add(80,"UI/UX Design","Positive","Loading state shown during authentication","Valid credentials","1. Submit login form\n2. Observe during processing","Visual loading indicator visible during auth","Medium")
add(81,"UI/UX Design","Edge Case","Login page on slow 3G network","Network throttled","1. Throttle to Slow 3G\n2. Load login page","Page loads with acceptable delay; no broken assets","Medium")
add(82,"UI/UX Design","Positive","Free trial signup CTA visible on login page","None","1. Navigate to login page\n2. Look for registration/trial CTA","Clear call-to-action for free trial signup visible","Medium")

# === MODULE 7: ACCESSIBILITY (TC_083 - TC_092) ===
add(83,"Accessibility","Positive","Full keyboard navigation on login page","None","1. Use Tab to navigate all interactive elements\n2. Use Enter to submit","All elements reachable and operable via keyboard","High")
add(84,"Accessibility","Positive","Screen reader reads all form labels (ARIA)","Screen reader active","1. Enable screen reader\n2. Navigate login form","All labels, inputs, buttons announced correctly","High")
add(85,"Accessibility","Positive","High-contrast mode support","None","1. Enable high-contrast mode in OS/browser\n2. Navigate login page","All text and controls visible in high-contrast","Medium")
add(86,"Accessibility","Positive","Focus indicator visible on all interactive elements","None","1. Tab through login page elements","Visible focus ring/outline on each focused element","High")
add(87,"Accessibility","Negative","Missing alt text on images","None","1. Inspect images on login page\n2. Check for alt attributes","All images have descriptive alt text","Medium")
add(88,"Accessibility","Positive","Error messages linked to fields via aria-describedby","Validation error triggered","1. Trigger validation error\n2. Inspect ARIA attributes","Error messages programmatically linked to their fields","Medium")
add(89,"Accessibility","Edge Case","Login page with screen magnifier at 400%","Magnifier active","1. Use screen magnifier at 400%\n2. Navigate login form","All elements accessible; no content lost","Low")
add(90,"Accessibility","Positive","Color contrast ratio meets WCAG 2.1 AA","None","1. Inspect text/background contrast ratios","All text meets minimum 4.5:1 contrast ratio","High")
add(91,"Accessibility","Negative","Form submission without filling required fields","None","1. Tab to Login button\n2. Press Enter with empty fields","Screen reader announces validation errors","Medium")
add(92,"Accessibility","Positive","Skip navigation link available","None","1. Load login page\n2. First Tab press","Skip to main content link is available","Low")

# === MODULE 8: SECURITY & COMPLIANCE (TC_093 - TC_100) ===
add(93,"Security","Positive","Login page served over HTTPS","None","1. Navigate to app.vwo.com\n2. Check URL and certificate","Page served via HTTPS with valid TLS certificate","Critical")
add(94,"Security","Positive","Password field is masked by default","None","1. Navigate to login page\n2. Enter password","Password characters are masked (dots/asterisks)","High")
add(95,"Security","Negative","Brute force attack protection","None","1. Attempt 20+ failed logins rapidly\n2. Observe system response","Account locked or CAPTCHA/rate limiting triggered","Critical")
add(96,"Security","Negative","Man-in-the-middle: HTTP downgrade attempt","None","1. Try accessing app.vwo.com via HTTP\n2. Observe response","Request redirected to HTTPS; no data sent over HTTP","Critical")
add(97,"Security","Positive","Password not visible in page source or network logs","None","1. Submit login\n2. Inspect Network tab for password in payload","Password transmitted securely; not in plain text in URL","Critical")
add(98,"Security","Edge Case","Login from a known compromised IP/VPN","Threat intelligence active","1. Login from flagged IP\n2. Observe response","Additional verification or blocking per security policy (Needs clarification)","Medium")
add(99,"Security","Positive","GDPR-compliant cookie consent on login page","EU region user","1. Navigate to login page from EU\n2. Observe cookie banner","Cookie consent banner displayed before non-essential cookies set","High")
add(100,"Security","Positive","Audit log captures login events","Admin access available","1. Login with valid credentials\n2. Check audit logs","Login event recorded with timestamp, IP, user agent","High")

# === WRITE TO EXCEL ===
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

headers = ["Test Case ID","Module","Category","Description","Pre-conditions","Steps","Expected Result","Priority"]
header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

# Priority colors
pri_fills = {
    "Critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "High": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Medium": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "Low": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
}
cat_fills = {
    "Positive": PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid"),
    "Negative": PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid"),
    "Edge Case": PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
    "Boundary": PatternFill(start_color="FDF4FF", end_color="FDF4FF", fill_type="solid"),
}

for i, tc in enumerate(TC, 2):
    vals = [tc["id"], tc["module"], tc["category"], tc["description"], tc["preconditions"], tc["steps"], tc["expected"], tc["priority"]]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=col, value=v)
        c.alignment = wrap
        c.border = border
        c.font = Font(name="Calibri", size=10)
    # Color priority cell
    ws.cell(row=i, column=8).fill = pri_fills.get(tc["priority"], PatternFill())
    # Color category cell
    ws.cell(row=i, column=3).fill = cat_fills.get(tc["category"], PatternFill())

# Column widths
widths = [12, 22, 14, 45, 30, 50, 45, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Freeze header
ws.freeze_panes = "A2"
# Auto-filter
ws.auto_filter.ref = f"A1:H{len(TC)+1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "VWO Login Dashboard - Test Case Summary"
ws2["A1"].font = Font(size=14, bold=True)
ws2.merge_cells("A1:C1")

summaries = [("",""),("Module","Count")]
modules = {}
for tc in TC:
    modules[tc["module"]] = modules.get(tc["module"], 0) + 1
for m, cnt in modules.items():
    summaries.append((m, cnt))
summaries.append(("TOTAL", len(TC)))
summaries.append(("",""))
summaries.append(("Category","Count"))
cats = {}
for tc in TC:
    cats[tc["category"]] = cats.get(tc["category"], 0) + 1
for ca, cnt in cats.items():
    summaries.append((ca, cnt))
summaries.append(("",""))
summaries.append(("Priority","Count"))
pris = {}
for tc in TC:
    pris[tc["priority"]] = pris.get(tc["priority"], 0) + 1
for p, cnt in pris.items():
    summaries.append((p, cnt))

for r, (a, b) in enumerate(summaries, 2):
    ws2.cell(row=r, column=1, value=a).font = Font(name="Calibri", size=10, bold=(a in ("Module","Category","Priority","TOTAL")))
    ws2.cell(row=r, column=2, value=b).font = Font(name="Calibri", size=10, bold=(a=="TOTAL"))
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 12

out = r"c:\Users\akank\OneDrive\Documents\All_abt_AI\RAG\Test_Cases_VWO_Login.xlsx"
wb.save(out)
print(f"SUCCESS: {len(TC)} test cases written to {out}")
